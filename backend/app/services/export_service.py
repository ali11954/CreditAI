import csv
import os
import tempfile
from io import BytesIO
from typing import Any, Optional

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)


async def export_to_excel(
    data: list[dict[str, Any]],
    filename: str,
    sheet_name: str = "Sheet1",
    columns: Optional[list[dict[str, str]]] = None,
) -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_alignment = Alignment(vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    if columns:
        if isinstance(columns[0], str):
            keys = columns
            labels = columns
        else:
            keys = [col["key"] for col in columns]
            labels = [col.get("label", col["key"]) for col in columns]
    elif data:
        keys = list(data[0].keys())
        labels = keys
    else:
        keys = []
        labels = []

    for col_idx, label in enumerate(labels, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    for row_idx, row_data in enumerate(data, 2):
        for col_idx, key in enumerate(keys, 1):
            value = row_data.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border

    for col in ws.columns:
        max_length = 0
        column_letter = col[0].column_letter
        for cell in col:
            try:
                cell_value = str(cell.value) if cell.value is not None else ""
                if len(cell_value) > max_length:
                    max_length = len(cell_value)
            except Exception:
                pass
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = max(adjusted_width, 12)

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, "wb") as f:
        f.write(output.read())

    return filename


async def export_to_pdf(
    data: list[dict[str, Any]],
    filename: str,
    title: str,
    columns: Optional[list[dict[str, str]]] = None,
    orientation: str = "landscape",
) -> str:
    pdfmetrics.registerFont(
        TTFont("Arabic", "C:/Windows/Fonts/arial.ttf")
    )

    page_size = landscape(A4) if orientation == "landscape" else A4

    output = BytesIO()
    doc = SimpleDocTemplate(
        output,
        pagesize=page_size,
        rightMargin=20,
        leftMargin=20,
        topMargin=30,
        bottomMargin=30,
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "CustomTitle",
        parent=styles["Heading1"],
        fontSize=18,
        fontName="Arabic",
        spaceAfter=20,
        alignment=1,
    )

    header_style = ParagraphStyle(
        "HeaderCell",
        parent=styles["Normal"],
        fontName="Arabic",
        fontSize=8,
        textColor=colors.white,
        leading=12,
    )

    cell_style = ParagraphStyle(
        "DataCell",
        parent=styles["Normal"],
        fontName="Arabic",
        fontSize=7,
        leading=10,
    )

    elements = []
    elements.append(Paragraph(title, title_style))
    elements.append(Spacer(1, 12))

    if not data:
        elements.append(Paragraph("No data available", cell_style))
        doc.build(elements)
        output.seek(0)
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        with open(filename, "wb") as f:
            f.write(output.read())
        return filename

    if columns:
        if isinstance(columns[0], str):
            keys = columns
            labels = columns
        else:
            keys = [col["key"] for col in columns]
            labels = [col.get("label", col["key"]) for col in columns]
    else:
        keys = list(data[0].keys())
        labels = keys

    header_row = [Paragraph(label, header_style) for label in labels]
    table_data = [header_row]

    for row_data in data:
        row = []
        for key in keys:
            value = row_data.get(key, "")
            row.append(Paragraph(str(value) if value is not None else "", cell_style))
        table_data.append(row)

    table = Table(table_data, repeatRows=1)
    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#002060")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, -1), "Arabic"),
                ("FONTSIZE", (0, 0), (-1, 0), 9),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
                ("TOPPADDING", (0, 0), (-1, 0), 10),
                ("BACKGROUND", (0, 1), (-1, -1), colors.HexColor("#F5F5F5")),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F5")]),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CCCCCC")),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("TOPPADDING", (0, 1), (-1, -1), 4),
                ("BOTTOMPADDING", (0, 1), (-1, -1), 4),
            ]
        )
    )
    elements.append(table)

    doc.build(elements)
    output.seek(0)

    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, "wb") as f:
        f.write(output.read())

    return filename


async def export_to_csv(
    data: list[dict[str, Any]],
    filename: str,
    columns: Optional[list[dict[str, str]]] = None,
) -> str:
    if not data:
        os.makedirs(os.path.dirname(filename), exist_ok=True)
        with open(filename, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            if columns:
                writer.writerow([col["label"] for col in columns])
        return filename

    if columns:
        if isinstance(columns[0], str):
            keys = columns
            labels = columns
        else:
            keys = [col["key"] for col in columns]
            labels = [col.get("label", col["key"]) for col in columns]
    else:
        keys = list(data[0].keys())
        labels = keys

    os.makedirs(os.path.dirname(filename), exist_ok=True)
    with open(filename, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.writer(f)
        writer.writerow(labels)
        for row_data in data:
            row = [row_data.get(key, "") for key in keys]
            writer.writerow(row)

    return filename
