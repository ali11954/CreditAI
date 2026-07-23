from typing import List, Dict, Any
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def create_excel_report(
    data: List[Dict[str, Any]],
    headers: List[str],
    title: str = "Report"
) -> BytesIO:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title
    
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="002060", end_color="002060", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def parse_excel_to_dict(file_content: bytes) -> List[Dict[str, Any]]:
    wb = openpyxl.load_workbook(BytesIO(file_content))
    ws = wb.active
    
    headers = [cell.value for cell in ws[1]]
    
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))
        data.append(row_dict)
    
    return data
